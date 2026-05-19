"""
upload.py — Excel Mass Upload endpoints.

Supported upload types:
  1. Inbound Serials  (against a PO)
  2. Warehouse State Update (bulk)
  3. Outbound Order Allocation (serials to an outbound order)

GET  /api/upload/template/{type}   — download the blank Excel template
POST /api/upload/inbound           — form: file + po_id
POST /api/upload/state-update      — form: file
POST /api/upload/outbound-alloc    — form: file + order_id
"""

import io
from datetime import datetime, timezone
from typing import Optional

import openpyxl
import xlrd
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models import (
    InboundShipment, OutboundOrder, OutboundOrderLine, OutboundOrderSerial,
    Product, PurchaseOrder, PurchaseOrderLine,
    SerialNumber, StateHistory, TerminalState, User,
)

router = APIRouter(prefix="/api/upload", tags=["upload"])

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

NAVY = "003087"
BLUE = "0075C2"
HEADER_FILL = openpyxl.styles.PatternFill("solid", fgColor=NAVY)
HEADER_FONT = openpyxl.styles.Font(bold=True, color="FFFFFF", name="Calibri")
REQUIRED_FILL = openpyxl.styles.PatternFill("solid", fgColor="E8F0FE")
NOTE_FONT = openpyxl.styles.Font(italic=True, color="888888", name="Calibri")


def _excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header_row(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _read_excel(file_bytes: bytes) -> list[dict]:
    """Read first sheet of an Excel file; return list of row dicts (header = row 1)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
               for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # skip blank rows
        result.append(dict(zip(headers, row)))
    return result


def _get_state(db: Session, code: str) -> Optional[TerminalState]:
    return db.query(TerminalState).filter(TerminalState.code == code).first()


def _record_state(db: Session, serial: SerialNumber, state: TerminalState,
                  user: User, notes: str = None):
    serial.current_state_id = state.id
    history = StateHistory(
        serial_number_id=serial.id,
        state_id=state.id,
        location_id=serial.current_location_id,
        datetime_utc=datetime.now(timezone.utc),
        timezone="UTC",
        actor_type="user",
        actor_user_id=user.id,
        notes=notes,
    )
    db.add(history)


# ---------------------------------------------------------------------------
# Template downloads
# ---------------------------------------------------------------------------

@router.get("/template/{upload_type}")
def download_template(
    upload_type: str,
    current_user: User = Depends(get_current_user),
):
    if upload_type == "inbound":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inbound Serials"
        headers = ["serial_number", "product_code"]
        notes   = ["Required – serial number from supplier", "Required – product code (must match master data)"]
        ws.append(headers)
        _style_header_row(ws, len(headers))
        ws.append(["", ""])  # example blank row
        # column widths
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 22
        # notes row
        for col, note in enumerate(notes, 1):
            c = ws.cell(row=3, column=col, value=note)
            c.font = NOTE_FONT
        return _excel_response(wb, "template_inbound_serials.xlsx")

    if upload_type == "state-update":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "State Update"
        headers = ["serial_number", "new_state", "notes"]
        notes   = [
            "Required – serial number",
            "Required – state CODE (e.g. AVAILABLE, STAGING, QUARANTINE …)",
            "Optional – reason / notes",
        ]
        ws.append(headers)
        _style_header_row(ws, len(headers))
        ws.append(["", "", ""])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 40
        for col, note in enumerate(notes, 1):
            c = ws.cell(row=3, column=col, value=note)
            c.font = NOTE_FONT
        return _excel_response(wb, "template_state_update.xlsx")

    if upload_type == "outbound-alloc":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outbound Allocation"
        headers = ["serial_number"]
        notes   = ["Required – serial number to allocate to the selected outbound order"]
        ws.append(headers)
        _style_header_row(ws, len(headers))
        ws.append([""])
        ws.column_dimensions["A"].width = 30
        for col, note in enumerate(notes, 1):
            c = ws.cell(row=3, column=col, value=note)
            c.font = NOTE_FONT
        return _excel_response(wb, "template_outbound_allocation.xlsx")

    if upload_type == "terminals-receiving":
        # Template matches Terminals_receiving-Sample.xls format
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lot-001"
        # Row 1-2: product header
        ws.cell(1, 1, "PN").font = HEADER_FONT
        ws.cell(1, 2, "Type").font = HEADER_FONT
        ws.cell(1, 3, "Total Qty")
        ws.cell(2, 1, "e.g. PAX-A920")
        ws.cell(2, 2, "PAX A920")
        ws.cell(2, 3, 10)
        # Row 3: blank separator
        # Row 4: column headers (matches sample file row 3, 1-indexed = row 4)
        headers = ["INDEX","Type","Part#","Lot#","Serial #","PO#","WIFI MAC","BT MAC","ETHERNET MAC","IMEI 1","IMEI 2","ICCID","EID","Key ID"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(4, ci, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(h) + 2, 12)
        # Row 5: example data row
        ws.cell(5, 1, 1); ws.cell(5, 2, "PAX A920"); ws.cell(5, 3, "PAX-A920")
        ws.cell(5, 4, "LOT-001"); ws.cell(5, 5, "SN12345678"); ws.cell(5, 6, "PO000001")
        ws.cell(5, 7, "AA:BB:CC:DD:EE:FF"); ws.cell(5, 8, "11:22:33:44:55:66")
        return _excel_response(wb, "template_terminals_receiving.xlsx")

    if upload_type == "products":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["code","name","product_type","product_category","serialised","unit_value","unit_currency","hs_code","active"]
        notes = ["Required","Required","Payment Terminal|Accessory|Battery","PaymentDevice|SerializedAccessory|Accessory","0 or 1","Numeric","e.g. EUR","HS code","0 or 1"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["PAX-A920","PAX A920 Terminal","Payment Terminal","PaymentDevice",1,450,"EUR","8470.50.00",1])
        for ci, note in enumerate(notes, 1):
            c = ws.cell(3+1, ci, note); c.font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 20) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_products.xlsx")

    if upload_type == "purchases":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Orders"
        headers = ["supplier_code","destination_location_code","order_date","expected_arrival_date","external_reference","notes"]
        notes = ["Required – supplier code","Required – location code","YYYY-MM-DD","YYYY-MM-DD","External ref/PO#","Optional notes"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["SUP001","WH-AMS","2026-06-01","2026-06-15","AU00000320",""])
        for ci, note in enumerate(notes, 1):
            ws.cell(3+1, ci, note).font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 24) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_purchase_orders.xlsx")

    if upload_type == "sales-orders":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Orders"
        headers = ["order_type","customer_ref","fulfilling_location_code","product_code","quantity"]
        notes = ["Sales|Rental|Replacement","Customer reference","Location code","Product code","Integer quantity"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["Sales","CUST001","WH-AMS","PAX-A920",5])
        for ci, note in enumerate(notes, 1):
            ws.cell(3+1, ci, note).font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 22) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_sales_orders.xlsx")

    if upload_type == "dist-outbound":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Distribution Outbound"
        headers = ["origin_location_code","destination_location_code","product_code","quantity","stock"]
        notes = ["Origin location code","Destination location code","Product code","Integer","Live|Refurbished"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["WH-AMS","WH-LHR","PAX-A920",10,"Live"])
        for ci, note in enumerate(notes, 1):
            ws.cell(3+1, ci, note).font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 24) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_distribution_outbound.xlsx")

    if upload_type == "dist-inbound":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Distribution Inbound"
        headers = ["distribution_order_number","serial_number","product_code","quantity","product_state"]
        notes = ["DS order number","Serial number (if serialised)","Product code","Integer","AVAILABLE|QUARANTINE_REFURBISHED"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["DS000001","SN12345678","PAX-A920",1,"AVAILABLE"])
        for ci, note in enumerate(notes, 1):
            ws.cell(3+1, ci, note).font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 26) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_distribution_inbound.xlsx")

    if upload_type == "rr-outbound":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RR Outbound Dispatch"
        headers = ["dispatch_type","serial_number","product_code","reason","location_code"]
        notes = ["Repair|Rework","Serial number","Product code","Reason for dispatch","Originating warehouse code"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["Repair","SN12345678","PAX-A920","Screen cracked","WH-AMS"])
        for ci, note in enumerate(notes, 1):
            ws.cell(3+1, ci, note).font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 22) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_rr_outbound.xlsx")

    if upload_type == "rr-inbound":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RR Inbound Return"
        headers = ["rr_order_number","serial_number","outcome","actual_cost","actual_cost_currency","repair_notes"]
        notes = ["RR order number","Serial number","Repaired|Beyond Repair","Cost amount","Currency code e.g. EUR","Optional notes"]
        ws.append(headers); _style_header_row(ws, len(headers))
        ws.append(["RR000001","SN12345678","Repaired",120,"EUR","Replaced screen"])
        for ci, note in enumerate(notes, 1):
            ws.cell(3+1, ci, note).font = NOTE_FONT
        [setattr(ws.column_dimensions[openpyxl.utils.get_column_letter(i)], 'width', 22) for i in range(1, len(headers)+1)]
        return _excel_response(wb, "template_rr_inbound.xlsx")

    raise HTTPException(status_code=404, detail=f"Unknown template type: {upload_type}")


# ---------------------------------------------------------------------------
# 1. Inbound Serial Import
# ---------------------------------------------------------------------------

@router.post("/inbound")
async def upload_inbound(
    po_id: int = Form(...),
    shipment_reference: Optional[str] = Form(None),
    carrier: Optional[str] = Form(None),
    carrier_tracking_ref: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
    if po.status not in ("Issued", "Partially Received"):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot import serials — PO status is '{po.status}' (must be Issued or Partially Received)",
        )

    raw = await file.read()
    rows = _read_excel(raw)

    expecting_state = _get_state(db, "EXPECTING")
    if not expecting_state:
        raise HTTPException(status_code=500, detail="EXPECTING state not found in database")

    # Build product lookup
    all_products = db.query(Product).filter(Product.active == 1).all()
    product_map = {p.code.upper(): p for p in all_products}

    # Build PO line lookup by product_id
    po_lines = {line.product_id: line for line in po.lines}

    created = 0
    duplicates = 0
    errors = []

    # Create a shipment record
    shipment = None
    if shipment_reference or carrier or carrier_tracking_ref:
        shipment = InboundShipment(
            po_id=po_id,
            shipment_reference=shipment_reference,
            carrier=carrier,
            carrier_tracking_ref=carrier_tracking_ref,
            uploaded_by_user_id=current_user.id,
        )
        db.add(shipment)
        db.flush()

    for i, row in enumerate(rows, start=2):
        sn_val = str(row.get("serial_number") or "").strip()
        pc_val = str(row.get("product_code") or "").strip().upper()

        if not sn_val or sn_val.lower() == "none":
            errors.append(f"Row {i}: serial_number is empty")
            continue
        if not pc_val or pc_val.lower() == "none":
            errors.append(f"Row {i}: product_code is empty")
            continue

        product = product_map.get(pc_val)
        if not product:
            errors.append(f"Row {i}: product code '{pc_val}' not found")
            continue

        # Check for duplicate
        existing = db.query(SerialNumber).filter(
            SerialNumber.serial_number == sn_val,
            SerialNumber.product_id == product.id,
        ).first()
        if existing:
            duplicates += 1
            continue

        # Get or create PO line for this product
        po_line = po_lines.get(product.id)
        po_line_id = po_line.id if po_line else None

        serial = SerialNumber(
            serial_number=sn_val,
            supplier_id=po.supplier_id,
            product_id=product.id,
            current_state_id=expecting_state.id,
            current_location_id=po.destination_location_id,
            stock_type="Live",
            po_id=po_id,
            po_line_id=po_line_id,
            active=1,
            accumulated_cost=0,
        )
        db.add(serial)
        db.flush()

        history = StateHistory(
            serial_number_id=serial.id,
            state_id=expecting_state.id,
            location_id=po.destination_location_id,
            datetime_utc=datetime.now(timezone.utc),
            timezone="UTC",
            actor_type="user",
            actor_user_id=current_user.id,
            notes=f"Imported via Excel upload — PO {po.po_number}",
        )
        db.add(history)

        # Update PO line qty_expected
        if po_line:
            po_line.qty_expected += 1

        created += 1

    # Update PO status
    if created > 0:
        all_received = all(
            line.qty_expected <= line.qty_received for line in po.lines
        )
        po.status = "Fully Received" if all_received else "Partially Received"

    db.commit()

    return {
        "total_rows": len(rows),
        "created": created,
        "duplicates": duplicates,
        "errors": errors,
        "po_number": po.po_number,
    }


# ---------------------------------------------------------------------------
# 2. Warehouse State Update
# ---------------------------------------------------------------------------

@router.post("/state-update")
async def upload_state_update(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    raw = await file.read()
    rows = _read_excel(raw)

    # Build state lookup by code
    all_states = db.query(TerminalState).filter(TerminalState.active == 1).all()
    state_map = {s.code.upper(): s for s in all_states}

    updated = 0
    not_found = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        sn_val    = str(row.get("serial_number") or "").strip()
        state_val = str(row.get("new_state") or "").strip().upper()
        notes_val = str(row.get("notes") or "").strip() or None

        if not sn_val or sn_val.lower() == "none":
            errors.append(f"Row {i}: serial_number is empty")
            continue
        if not state_val or state_val.lower() == "none":
            errors.append(f"Row {i}: new_state is empty")
            continue

        state = state_map.get(state_val)
        if not state:
            errors.append(f"Row {i}: state code '{state_val}' not recognised")
            continue

        serial = db.query(SerialNumber).filter(
            SerialNumber.serial_number == sn_val,
            SerialNumber.active == 1,
        ).first()
        if not serial:
            not_found += 1
            errors.append(f"Row {i}: serial '{sn_val}' not found")
            continue

        _record_state(db, serial, state, current_user, notes=notes_val)
        updated += 1

    db.commit()

    return {
        "total_rows": len(rows),
        "updated": updated,
        "not_found": not_found,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# 3. Outbound Order Allocation
# ---------------------------------------------------------------------------

@router.post("/outbound-alloc")
async def upload_outbound_alloc(
    order_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    order = db.query(OutboundOrder).filter(OutboundOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Outbound Order not found")
    if order.status not in ("Issued", "Allocated"):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot allocate — order status is '{order.status}' (must be Issued or Allocated)",
        )

    raw = await file.read()
    rows = _read_excel(raw)

    available_state = _get_state(db, "AVAILABLE")
    if not available_state:
        raise HTTPException(status_code=500, detail="AVAILABLE state not found")

    # Build a quick line map: product_id → order_line
    line_map = {line.product_id: line for line in order.lines}

    # Serials already on this order
    already_allocated = {os.serial_id for os in order.serials}

    allocated = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        sn_val = str(row.get("serial_number") or "").strip()
        if not sn_val or sn_val.lower() == "none":
            errors.append(f"Row {i}: serial_number is empty")
            continue

        serial = db.query(SerialNumber).filter(
            SerialNumber.serial_number == sn_val,
            SerialNumber.active == 1,
        ).first()
        if not serial:
            errors.append(f"Row {i}: serial '{sn_val}' not found")
            continue

        if serial.id in already_allocated:
            skipped += 1
            continue

        # Find matching order line
        line = line_map.get(serial.product_id)
        if not line:
            errors.append(f"Row {i}: serial '{sn_val}' product not on this order")
            continue

        alloc = OutboundOrderSerial(
            order_id=order.id,
            order_line_id=line.id,
            serial_id=serial.id,
        )
        db.add(alloc)
        already_allocated.add(serial.id)
        allocated += 1

    if allocated > 0 and order.status == "Issued":
        order.status = "Allocated"

    db.commit()

    return {
        "total_rows": len(rows),
        "allocated": allocated,
        "skipped_duplicates": skipped,
        "errors": errors,
        "order_number": order.order_number,
    }


# ---------------------------------------------------------------------------
# 4. Terminals Receiving — multi-sheet XLS (Terminals_receiving-Sample format)
# ---------------------------------------------------------------------------

def _read_xls_multi_sheet(raw: bytes) -> list[dict]:
    """
    Parse multi-sheet XLS in the Terminals_receiving-Sample format.
    Each sheet: row 0-1 = product header, row 2 = blank, row 3 = column headers, rows 4+ = data.
    Returns flat list of row dicts with added 'sheet_name' key.
    """
    wb = xlrd.open_workbook(file_contents=raw)
    results = []
    for sheet in wb.sheets():
        nrows = sheet.nrows
        if nrows < 5:
            continue
        # Row 3 (0-indexed) = column headers
        raw_headers = sheet.row_values(3)
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(raw_headers)]
        # normalise header names for easy mapping
        norm = {h.lower().replace(" ", "_").replace("#", "num").replace("/", "_"): h for h in headers}
        header_map = {orig: i for i, orig in enumerate(headers)}

        for rn in range(4, nrows):
            row_vals = sheet.row_values(rn)
            if all(v == "" or v is None for v in row_vals):
                continue
            row_dict = {"sheet_name": sheet.name}
            for h, ci in header_map.items():
                val = row_vals[ci] if ci < len(row_vals) else None
                # xlrd returns floats for numbers — convert INDEX to int
                if h.upper() == "INDEX" and isinstance(val, float):
                    val = int(val)
                row_dict[h] = val
            results.append(row_dict)
    return results


@router.post("/terminals-receiving")
async def upload_terminals_receiving(
    po_id: Optional[int] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Upload multi-sheet XLS in Terminals_receiving-Sample format.
    Columns: INDEX, Type, Part#, Lot#, Serial #, PO#, WIFI MAC, BT MAC, ETHERNET MAC, IMEI 1, IMEI 2, ICCID, EID, Key ID
    Creates SerialNumber records in EXPECTING state.
    If po_id is not provided, matches from file's PO# column.
    """
    raw = await file.read()

    # Try XLS first, fall back to XLSX
    try:
        rows = _read_xls_multi_sheet(raw)
    except Exception:
        # Fall back to xlsx parser if xlrd fails (e.g. file is actually xlsx)
        rows = _read_excel(raw)
        # Map xlsx headers to expected format
        remapped = []
        for r in rows:
            remapped.append({
                "Serial #": r.get("serial_#") or r.get("serial_number") or "",
                "Part#": r.get("partnum") or r.get("part_num") or r.get("product_code") or "",
                "Lot#": r.get("lot_num") or r.get("lot_number") or r.get("lotnum") or "",
                "PO#": r.get("po_num") or r.get("ponum") or r.get("po_number") or "",
                "Type": r.get("type") or r.get("terminal_type") or "",
                "WIFI MAC": r.get("wifi_mac") or r.get("wifi mac") or "",
                "BT MAC": r.get("bt_mac") or r.get("bluetooth_mac") or "",
                "ETHERNET MAC": r.get("ethernet_mac") or "",
                "IMEI 1": r.get("imei_1") or r.get("imei1") or "",
                "IMEI 2": r.get("imei_2") or r.get("imei2") or "",
                "ICCID": r.get("iccid") or "",
                "EID": r.get("eid") or "",
                "Key ID": r.get("key_id") or r.get("keyid") or "",
            })
        rows = remapped

    expecting_state = _get_state(db, "EXPECTING")
    if not expecting_state:
        raise HTTPException(status_code=500, detail="EXPECTING state not found in database")

    all_products = db.query(Product).filter(Product.active == 1).all()
    product_by_code = {p.code.upper(): p for p in all_products}

    # PO lookup if po_id provided
    po_obj = None
    if po_id:
        po_obj = db.query(PurchaseOrder).get(po_id)

    # Build PO lookup by po_number for file-driven matching
    all_pos = db.query(PurchaseOrder).all()
    po_by_number = {p.po_number.upper(): p for p in all_pos if p.po_number}

    created = 0
    duplicates = 0
    errors = []

    for i, row in enumerate(rows, start=1):
        sn_val = str(row.get("Serial #") or row.get("serial_number") or "").strip()
        pc_val = str(row.get("Part#") or row.get("product_code") or "").strip().upper()

        if not sn_val or sn_val.lower() in ("none", ""):
            errors.append(f"Row {i}: serial number is empty — skipped")
            continue
        if not pc_val or pc_val.lower() in ("none", ""):
            errors.append(f"Row {i}: product code (Part#) is empty — skipped")
            continue

        product = product_by_code.get(pc_val)
        if not product:
            errors.append(f"Row {i}: product '{pc_val}' not found in master data")
            continue

        # Resolve PO
        row_po_num = str(row.get("PO#") or "").strip().upper()
        resolved_po = po_obj
        if not resolved_po and row_po_num:
            resolved_po = po_by_number.get(row_po_num)

        # Check uniqueness
        existing = db.query(SerialNumber).filter(
            SerialNumber.serial_number == sn_val,
            SerialNumber.product_id == product.id,
        ).first()
        if existing:
            duplicates += 1
            continue

        serial = SerialNumber(
            serial_number=sn_val,
            supplier_id=resolved_po.supplier_id if resolved_po else product.id,  # fallback
            product_id=product.id,
            lot_number=str(row.get("Lot#") or "").strip() or None,
            terminal_type=str(row.get("Type") or "").strip() or None,
            wifi_mac=str(row.get("WIFI MAC") or "").strip() or None,
            bluetooth_mac=str(row.get("BT MAC") or "").strip() or None,
            ethernet_mac=str(row.get("ETHERNET MAC") or "").strip() or None,
            imei1=str(row.get("IMEI 1") or "").strip() or None,
            imei2=str(row.get("IMEI 2") or "").strip() or None,
            iccid=str(row.get("ICCID") or "").strip() or None,
            eid=str(row.get("EID") or "").strip() or None,
            key_id=str(row.get("Key ID") or "").strip() or None,
            current_state_id=expecting_state.id,
            current_location_id=resolved_po.destination_location_id if resolved_po else None,
            stock_type="Live",
            po_id=resolved_po.id if resolved_po else None,
            active=1,
            accumulated_cost=0,
        )
        db.add(serial)
        db.flush()

        history = StateHistory(
            serial_number_id=serial.id,
            state_id=expecting_state.id,
            location_id=serial.current_location_id,
            datetime_utc=datetime.now(timezone.utc),
            timezone="UTC",
            actor_type="user",
            actor_user_id=current_user.id,
            activity_description=f"Terminal received via XLS upload (lot: {serial.lot_number or 'N/A'})",
            order_reference=resolved_po.po_number if resolved_po else None,
            notes=f"Imported from XLS — sheet: {row.get('sheet_name', 'N/A')}",
        )
        db.add(history)

        # Update PO line if resolved
        if resolved_po:
            po_line = next((l for l in resolved_po.lines if l.product_id == product.id), None)
            if po_line:
                po_line.qty_expected = (po_line.qty_expected or 0) + 1

        created += 1

    if created > 0 and po_obj:
        po_obj.status = "Expected"

    db.commit()

    return {
        "total_rows": len(rows),
        "created": created,
        "duplicates": duplicates,
        "errors": errors,
    }
